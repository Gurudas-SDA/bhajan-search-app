#!/usr/bin/env python3
"""
Merges validated translation batches into Bhajans.xlsx.

Steps:
  1. Backup original xlsx to scratchpad/Bhajans_backup_pre_translations.xlsx
  2. Load and merge all translation batches
  3. Add headers Spanish/Italian/French to columns I/J/K
  4. Write translations to xlsx
  5. Save and verify integrity
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata
import shutil
from pathlib import Path
import sys

def normalize_title(title):
    """Normalize title to NFC form."""
    if title is None:
        return None
    return unicodedata.normalize('NFC', str(title))

def load_batches(batch_dir):
    """Load all batch files and merge them."""
    batch_files = sorted(Path(batch_dir).glob('batch_*.json'))
    merged = {}

    for batch_file in batch_files:
        with open(batch_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        for item in data:
            title = normalize_title(item['title'])
            for verse in item.get('verses', []):
                verse_num = verse['number']
                key = (title, verse_num)
                merged[key] = {
                    'spanish': verse.get('spanish', ''),
                    'italian': verse.get('italian', ''),
                    'french': verse.get('french', ''),
                }

    return merged

def merge_translations():
    """Main merge function."""
    script_dir = Path(__file__).parent
    repo_root = script_dir.parent
    batch_dir = repo_root.parent / 'translations'
    xlsx_path = repo_root / 'Bhajans.xlsx'
    backup_path = repo_root.parent / 'Bhajans_backup_pre_translations.xlsx'

    print("=" * 70)
    print("TRANSLATION MERGE")
    print("=" * 70)

    # Step 1: Backup
    print("\n[1/5] Creating backup...")
    shutil.copy2(xlsx_path, backup_path)
    print(f"  Backed up to: {backup_path.name}")

    # Step 2: Load translations
    print("[2/5] Loading translation batches...")
    translations = load_batches(batch_dir)
    print(f"  Loaded {len(translations)} translated verses")

    # Step 3: Open xlsx
    print("[3/5] Opening Bhajans.xlsx...")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Lapa1']

    # Verify pre-merge state: I/J/K should be empty
    print("[4/5] Adding headers and merging translations...")

    # Get header style from H1 if it exists
    header_cell_h = ws.cell(1, 8)
    header_style = None
    if header_cell_h.font or header_cell_h.fill:
        header_style = {
            'font': Font(bold=header_cell_h.font.bold if header_cell_h.font else False,
                        color=header_cell_h.font.color if header_cell_h.font else None),
            'fill': PatternFill(fill_type=header_cell_h.fill.fill_type if header_cell_h.fill else None,
                               start_color=header_cell_h.fill.start_color if header_cell_h.fill else None),
        }

    # Add headers
    ws.cell(1, 9).value = 'Spanish'
    ws.cell(1, 10).value = 'Italian'
    ws.cell(1, 11).value = 'French'

    # Apply header style if it exists
    if header_style:
        for col in [9, 10, 11]:
            cell = ws.cell(1, col)
            if header_style['font']:
                cell.font = header_style['font']
            if header_style['fill']:
                cell.fill = header_style['fill']

    # Write translations
    rows_written = {9: 0, 10: 0, 11: 0}
    total_rows = 0

    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, 2).value
        verse_num = ws.cell(row, 4).value
        english = ws.cell(row, 6).value

        if title is None or verse_num is None:
            continue

        title_nfc = normalize_title(title)
        key = (title_nfc, int(verse_num))
        total_rows += 1

        if key in translations:
            trans = translations[key]
            ws.cell(row, 9).value = trans['spanish']
            ws.cell(row, 10).value = trans['italian']
            ws.cell(row, 11).value = trans['french']

            # Copy wrap_text style from column H if present
            h_cell = ws.cell(row, 8)
            if h_cell.alignment and h_cell.alignment.wrap_text:
                for col in [9, 10, 11]:
                    i_cell = ws.cell(row, col)
                    if i_cell.alignment:
                        i_cell.alignment = Alignment(wrap_text=True,
                                                    horizontal=i_cell.alignment.horizontal,
                                                    vertical=i_cell.alignment.vertical)
                    else:
                        i_cell.alignment = Alignment(wrap_text=True)

            # Count rows with translations
            if trans['spanish']:
                rows_written[9] += 1
            if trans['italian']:
                rows_written[10] += 1
            if trans['french']:
                rows_written[11] += 1

    # Save
    print("[5/5] Saving and verifying...")
    wb.save(xlsx_path)
    print(f"  Saved {xlsx_path.name}")

    # Verification
    print("\nVerification:")
    wb_verify = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_verify = wb_verify['Lapa1']

    # Check row count
    if ws_verify.max_row == 1713:  # 1 header + 1712 data
        print(f"  ✓ Row count: {ws_verify.max_row} (1 header + 1712 data)")
    else:
        print(f"  ✗ Row count mismatch: {ws_verify.max_row} (expected 1713)")
        return 1

    # Check A-H integrity against backup
    print("  Checking columns A-H integrity...")
    wb_backup = openpyxl.load_workbook(backup_path, data_only=True)
    ws_backup = wb_backup['Lapa1']

    integrity_ok = True
    for row in range(2, min(ws_verify.max_row + 1, ws_backup.max_row + 1)):
        for col in range(1, 9):
            new_val = ws_verify.cell(row, col).value
            old_val = ws_backup.cell(row, col).value
            if new_val != old_val:
                print(f"  ✗ Row {row} Col {chr(64+col)}: changed from {old_val} to {new_val}")
                integrity_ok = False
                break
        if not integrity_ok:
            break

    if integrity_ok:
        print(f"  ✓ Columns A-H unchanged")
    else:
        print(f"  ✗ Integrity check failed")
        return 1

    # Check translation columns
    print("  Translation columns written:")
    print(f"    Column I (Spanish): {rows_written[9]} rows")
    print(f"    Column J (Italian): {rows_written[10]} rows")
    print(f"    Column K (French):  {rows_written[11]} rows")

    # Spot check
    print("\n  Spot checks:")
    for row in range(2, min(2 + 5, ws_verify.max_row + 1)):
        title = ws_verify.cell(row, 2).value
        spanish = ws_verify.cell(row, 9).value
        if spanish:
            print(f"    Row {row} '{title}' -> Spanish: {spanish[:50]}...")

    print("\n" + "=" * 70)
    print("Result: SUCCESS ✓")
    print("=" * 70)
    return 0

if __name__ == '__main__':
    exit_code = merge_translations()
    sys.exit(exit_code)
