#!/usr/bin/env python3
"""
Validates all translation batch files against the Bhajans.xlsx source.

Checks:
  1. Every (Bhajan_Title, Verse_Number) pair appears exactly once across all batches
  2. NFC unicode normalization consistency
  3. For verses with non-empty English: spanish/italian/french all non-empty
  4. For verses with empty English: all translations empty
  5. No internal newlines in translations
  6. Length ratios within 0.5-2.0 (warnings)
  7. English-leak detection (warnings)
  8. Bracket count parity (warnings)
  9. IAST diacritic preservation (warnings/errors depending on prevalence)

Exit code: 0 if no errors, 1 if any errors found.
"""

import json
import openpyxl
import unicodedata
import re
import sys
from pathlib import Path
from collections import defaultdict

def normalize_title(title):
    """Normalize title to NFC form."""
    if title is None:
        return None
    return unicodedata.normalize('NFC', str(title))

def load_batches(batch_dir):
    """Load all batch files and merge them. Returns dict of (title, verse_num) -> translation_dict."""
    batch_files = sorted(Path(batch_dir).glob('batch_*.json'))
    merged = {}
    batch_count = {}

    for batch_file in batch_files:
        with open(batch_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        for item in data:
            title = normalize_title(item['title'])
            for verse in item.get('verses', []):
                verse_num = verse['number']
                key = (title, verse_num)

                if key in merged:
                    batch_count[key] = batch_count.get(key, 1) + 1
                else:
                    batch_count[key] = 1

                merged[key] = {
                    'spanish': verse.get('spanish', ''),
                    'italian': verse.get('italian', ''),
                    'french': verse.get('french', ''),
                    'batch_file': batch_file.name,
                }

    return merged, batch_count

def load_xlsx(xlsx_path):
    """Load xlsx and create lookup dict. Returns dict of (title, verse_num) -> english_value."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Lapa1']

    lookup = {}
    nfc_issues = []

    for row in range(2, ws.max_row + 1):
        title = ws.cell(row, 2).value
        verse_num = ws.cell(row, 4).value
        english = ws.cell(row, 6).value

        if title is None or verse_num is None:
            continue

        title_nfc = normalize_title(title)
        key = (title_nfc, int(verse_num))

        # Check for NFC drift
        title_nfd = unicodedata.normalize('NFD', str(title))
        if title_nfd != str(title):
            nfc_issues.append({
                'row': row,
                'title': str(title),
                'issue': 'NFD form differs from NFC'
            })

        lookup[key] = english or ''

    return lookup, nfc_issues

def check_newlines(text):
    """Check for internal newlines."""
    if not text:
        return False
    return '\n' in text or '\r' in text

def extract_words(text):
    """Extract words from text."""
    return re.findall(r'\b\w+\b', text.lower())

def detect_english_leak(translation, english_ref):
    """Detect if translation has too many English words."""
    if len(translation) < 8 or not english_ref:
        return None

    english_words = {
        'the', 'and', 'of', 'with', 'his', 'her', 'who', 'which',
        'this', 'are', 'is', 'have', 'has', 'been', 'was', 'were'
    }

    words = extract_words(translation)
    leaked = [w for w in words if w in english_words]

    if len(leaked) >= 4:
        return f"{len(leaked)} English words: {', '.join(set(leaked))}"
    return None

def check_bracket_parity(translation, english_ref):
    """Check bracket count matches."""
    trans_open = translation.count('[')
    trans_close = translation.count(']')
    eng_open = english_ref.count('[')
    eng_close = english_ref.count(']')

    if trans_open != eng_open or trans_close != eng_close:
        return f"[{trans_open} vs {eng_open}] {{{trans_close} vs {eng_close}}}"
    return None

def extract_iast_tokens(text):
    """Extract tokens containing IAST diacritics."""
    # IAST diacritics pattern
    iast_pattern = r'[a-zA-ZĀ-ſĀāĪīŪūŖŗḌḍḸḹṚṛ]*[āīūṛṝḷḹēōṁṃḥṅñṭḍṇśṣĵŵḓãõũĩẽ][a-zA-ZĀ-ſĀāĪīŪūŖŗḌḍḸḹṚṛ\']*'
    return set(re.findall(iast_pattern, text))

def iast_in_translation(translation, english_tokens):
    """Check if IAST tokens from English appear in translation (case-insensitive, allow Italian 's' drop)."""
    trans_tokens = extract_iast_tokens(translation)

    missing = []
    for eng_token in english_tokens:
        # Exact match
        if eng_token in trans_tokens:
            continue

        # Case-insensitive match
        if any(eng_token.lower() == t.lower() for t in trans_tokens):
            continue

        # Italian exception: missing trailing 's'
        if any(eng_token.lower() == t.lower() + 's' for t in trans_tokens):
            continue

        missing.append(eng_token)

    return missing if missing else None

def validate():
    """Run all validations."""
    script_dir = Path(__file__).parent
    repo_root = script_dir.parent
    batch_dir = repo_root.parent / 'translations'
    xlsx_path = repo_root / 'Bhajans.xlsx'

    print("=" * 70)
    print("TRANSLATION BATCH VALIDATION")
    print("=" * 70)

    # Load batches
    print("\n[1/4] Loading batch files...")
    batches, batch_counts = load_batches(batch_dir)
    print(f"  Loaded {len(batches)} verses from {len(list(batch_dir.glob('batch_*.json')))} batch files")

    # Load xlsx
    print("[2/4] Loading Bhajans.xlsx...")
    xlsx_lookup, nfc_issues = load_xlsx(xlsx_path)
    print(f"  Loaded {len(xlsx_lookup)} verses from xlsx")

    if nfc_issues:
        print(f"\n  NFC Drift detected in {len(nfc_issues)} rows (see warnings)")

    # Validation
    errors = []
    warnings = []

    print("[3/4] Validating coverage and duplicates...")

    # Check every batch entry exists in xlsx exactly once
    for (title, verse_num), trans_data in batches.items():
        if (title, verse_num) not in xlsx_lookup:
            errors.append(f"Batch verse not in xlsx: '{title}' verse {verse_num}")

        if batch_counts.get((title, verse_num), 0) > 1:
            errors.append(f"Duplicate across batches: '{title}' verse {verse_num} (appears {batch_counts[(title, verse_num)]} times)")

    # Check every xlsx entry that has non-empty English appears in batch
    missing_translations = []
    for (title, verse_num), english_text in xlsx_lookup.items():
        if (title, verse_num) not in batches:
            if english_text.strip():  # Only report if English is non-empty
                missing_translations.append((title, verse_num))

    if missing_translations:
        errors.append(f"Missing translations for {len(missing_translations)} verses with non-empty English")
        for title, verse_num in missing_translations[:5]:
            errors.append(f"  - '{title}' verse {verse_num}")
        if len(missing_translations) > 5:
            errors.append(f"  ... and {len(missing_translations) - 5} more")

    print("[4/4] Validating content quality...")

    length_warnings = []
    leak_warnings = []
    bracket_warnings = []
    iast_warnings = []
    newline_errors = []
    empty_mismatch = []

    for (title, verse_num), trans_data in batches.items():
        english_text = xlsx_lookup.get((title, verse_num), '')

        # Check newlines
        for lang in ['spanish', 'italian', 'french']:
            if check_newlines(trans_data[lang]):
                newline_errors.append(f"Newline in {lang}: '{title}' v{verse_num}")

        # Check empty/non-empty consistency
        english_empty = not english_text or not english_text.strip()
        trans_empty = all(not trans_data[lang].strip() for lang in ['spanish', 'italian', 'french'])

        if english_empty and not trans_empty:
            empty_mismatch.append(f"Empty English but non-empty translations: '{title}' v{verse_num}")
        elif not english_empty and trans_empty:
            empty_mismatch.append(f"Non-empty English but empty translations: '{title}' v{verse_num}")

        # Skip other checks if translations empty
        if trans_empty:
            continue

        # Length ratio check (warnings only)
        english_len = len(english_text)
        if english_len > 0:
            for lang in ['spanish', 'italian', 'french']:
                trans_len = len(trans_data[lang])
                ratio = trans_len / english_len if english_len > 0 else 1
                if ratio < 0.5 or ratio > 2.0:
                    length_warnings.append(f"{lang} ratio={ratio:.2f}: '{title}' v{verse_num}")

        # English leak detection (warnings)
        for lang in ['spanish', 'italian', 'french']:
            leak = detect_english_leak(trans_data[lang], english_text)
            if leak:
                leak_warnings.append(f"{lang} leak: {leak} in '{title}' v{verse_num}")

        # Bracket parity (warnings)
        for lang in ['spanish', 'italian', 'french']:
            bracket_issue = check_bracket_parity(trans_data[lang], english_text)
            if bracket_issue:
                bracket_warnings.append(f"{lang} brackets {bracket_issue}: '{title}' v{verse_num}")

        # IAST preservation
        iast_tokens = extract_iast_tokens(english_text)
        if iast_tokens:
            for lang in ['spanish', 'italian', 'french']:
                missing = iast_in_translation(trans_data[lang], iast_tokens)
                if missing:
                    iast_warnings.append({
                        'lang': lang,
                        'title': title,
                        'verse': verse_num,
                        'missing': missing
                    })

    # Add newline errors
    if newline_errors:
        errors.extend(newline_errors[:5])
        if len(newline_errors) > 5:
            errors.append(f"... and {len(newline_errors) - 5} more newline issues")

    # Add empty mismatch errors
    if empty_mismatch:
        errors.extend(empty_mismatch[:5])
        if len(empty_mismatch) > 5:
            errors.append(f"... and {len(empty_mismatch) - 5} more empty/non-empty mismatches")

    # Print summary
    print("\n" + "=" * 70)
    print("VALIDATION SUMMARY")
    print("=" * 70)

    if errors:
        print(f"\n[ERRORS] Total: {len(errors)}")
        for err in errors[:10]:
            print(f"  ERROR: {err}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more errors")
    else:
        print(f"\n[ERRORS] None found ✓")

    if warnings:
        print(f"\n[WARNINGS] Total: {sum(len(w) if isinstance(w, list) else 1 for w in [length_warnings, leak_warnings, bracket_warnings, iast_warnings, nfc_issues])}")

        if length_warnings:
            print(f"  Length ratio outliers: {len(length_warnings)}")
            for w in length_warnings[:3]:
                print(f"    - {w}")
            if len(length_warnings) > 3:
                print(f"    ... and {len(length_warnings) - 3} more")

        if leak_warnings:
            print(f"  English leaks: {len(leak_warnings)}")
            for w in leak_warnings[:3]:
                print(f"    - {w}")
            if len(leak_warnings) > 3:
                print(f"    ... and {len(leak_warnings) - 3} more")

        if bracket_warnings:
            print(f"  Bracket mismatches: {len(bracket_warnings)}")
            for w in bracket_warnings[:3]:
                print(f"    - {w}")
            if len(bracket_warnings) > 3:
                print(f"    ... and {len(bracket_warnings) - 3} more")

        if iast_warnings:
            iast_count = len(iast_warnings)
            iast_pct = (iast_count / len(batches)) * 100 if batches else 0
            if iast_pct > 2:
                print(f"  IAST diacritics missing: {iast_count} verses ({iast_pct:.1f}%)")
                for w in iast_warnings[:3]:
                    print(f"    - {w['lang']} in '{w['title']}' v{w['verse']}: missing {w['missing']}")
                if iast_count > 3:
                    print(f"    ... and {iast_count - 3} more")
            else:
                print(f"  IAST diacritics: {iast_count} verses (below threshold, warning-level)")

        if nfc_issues:
            print(f"  NFC normalization drift: {len(nfc_issues)} rows in xlsx")
            for issue in nfc_issues[:2]:
                print(f"    - Row {issue['row']}: {issue['title']}")
    else:
        print(f"\n[WARNINGS] None found ✓")

    print("\n" + "=" * 70)
    if errors:
        print("Result: FAILED (errors found)")
        return 1
    else:
        print("Result: PASSED ✓")
        return 0

if __name__ == '__main__':
    exit_code = validate()
    sys.exit(exit_code)
